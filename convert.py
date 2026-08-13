r"""
Solve Analytics Pipeline
-------------------------
Parses a cstimer.net export (.txt), computes rolling performance metrics
(ao5 / ao12 / ao100, consistency stddev), and appends new solves + a
formatted summary block to an existing Excel workbook.

Usage:
    python convert.py                     # uses config.json / defaults
    python convert.py --dry-run           # parse + compute, don't write
    python convert.py --downloads-dir "D:\Downloads" --excel-path "C:\Cubing.xlsx"

Config:
    Optional config.json next to this script can fill in the actual paths for the following keys:
    {
        "downloads_dir": PATH_TO_DOWNLOADS_DIR,
        "excel_path": "EXCEL_PATH",
        "sheet_name": "ACTUAL_NAME",
        "backup_dir": "BACKUP_DIR_PATH",
    }
"""

import argparse
import ast
import glob
import json
import logging
import os
import shutil
import statistics
from dataclasses import dataclass
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("solve_pipeline")

ALIGNMENT = Alignment(horizontal="center", vertical="center")


# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #

@dataclass
class Config:
    downloads_dir: str
    excel_path: str
    sheet_name: str
    backup_dir: Optional[str] = None
    max_backups: int = 5


def load_config(args: argparse.Namespace) -> Config:
    """Merge config.json (if present) with CLI overrides. CLI wins."""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    file_cfg = {}
    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            file_cfg = json.load(f)
            log.info(f"Loaded config from {config_path}")

    def pick(cli_val, key, default=None):
        if cli_val is not None:
            return cli_val
        return file_cfg.get(key, default)

    cfg = Config(
        downloads_dir=pick(args.downloads_dir, "downloads_dir", r"C:\Users\tejpa\Downloads"),
        excel_path=pick(args.excel_path, "excel_path", r"C:\Users\tejpa\OneDrive\tej\OneDrive\Cubing.xlsx"),
        sheet_name=pick(args.sheet_name, "sheet_name", "MAIN(Start-08-31-2025-Sun)"),
        backup_dir=pick(args.backup_dir, "backup_dir", None),
        max_backups=file_cfg.get("max_backups", 5),
    )
    return cfg


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

def find_latest_cstimer_export(downloads_dir: str) -> str:
    pattern = os.path.join(downloads_dir, "cstimer*.txt")
    matches = glob.glob(pattern)
    if not matches:
        raise FileNotFoundError(f"No cstimer*.txt export found in {downloads_dir}")
    latest = max(matches, key=os.path.getmtime)
    log.info(f"Using export: {latest}")
    return latest


def extract_session_array(raw: str, session_key: str) -> str:
    """Pulls out just the `"session_key":[ ... ]` array text via bracket
    matching (quote-aware), instead of assuming a fixed byte offset around
    neighboring keys. Works regardless of what other sessions/keys exist,
    their order, or how cstimer formats the surrounding JSON."""
    marker = f'"{session_key}"'
    marker_idx = raw.find(marker)
    if marker_idx == -1:
        raise KeyError(f"'{session_key}' not found in export.")

    start = raw.find("[", marker_idx)
    if start == -1:
        raise ValueError(f"No array found after '{session_key}' key.")

    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(raw)):
        c = raw[i]
        if in_string:
            if escape:
                escape = False
            elif c == "\\":
                escape = True
            elif c == '"':
                in_string = False
            continue
        if c == '"':
            in_string = True
        elif c == "[":
            depth += 1
        elif c == "]":
            depth -= 1
            if depth == 0:
                return raw[start:i + 1]

    raise ValueError(f"Unbalanced brackets while extracting '{session_key}' array.")


def parse_cstimer_export(path: str, session_key: str = "session1") -> List[Tuple[int, float, str]]:
    """Returns list of (solve_number, time_seconds, pll) for the given session."""
    with open(path, "r") as f:
        raw = f.read()

    array_str = extract_session_array(raw, session_key)

    try:
        solves_raw = json.loads(array_str)
    except json.JSONDecodeError:
        try:
            solves_raw = ast.literal_eval(array_str)
        except (ValueError, SyntaxError) as e:
            raise ValueError(f"Could not parse '{session_key}' array as JSON or a Python literal: {e}")

    solves = []
    pll_counts = {}
    for i, solve in enumerate(solves_raw, start=1):
        time_ms = solve[0][1]
        pll = solve[2] if len(solve) > 2 else ""
        time_sec = int(time_ms) / 1000.0
        solves.append((i, time_sec, pll))
        if pll:
            pll_counts[pll] = pll_counts.get(pll, 0) + 1

    log.info(f"Parsed {len(solves)} solves. Distinct PLLs seen: {len(pll_counts)} "
              f"(total PLL-tagged solves: {sum(pll_counts.values())})")
    return solves


# --------------------------------------------------------------------------- #
# Rolling performance metrics
# --------------------------------------------------------------------------- #

def trimmed_average(window: List[float]) -> Optional[float]:
    """WCA-style average: drop best/worst ~5% (min 1), average the rest.
    DNFs aren't modeled here since cstimer times are already in seconds."""
    n = len(window)
    if n < 3:
        return None
    trim = max(1, round(n * 0.05))
    trimmed = sorted(window)[trim:n - trim]
    if not trimmed:
        return None
    return sum(trimmed) / len(trimmed)


def rolling_metrics(all_times: List[float], upto_index: int) -> dict:
    """Computes ao5 / ao12 / ao100 / stdev100 ending at solve index `upto_index`
    (1-indexed, inclusive) using the full time history."""
    window = all_times[:upto_index]

    def avg_last(n):
        return round(trimmed_average(window[-n:]), 3) if len(window) >= n else None

    stdev100 = None
    if len(window) >= 100:
        stdev100 = round(statistics.pstdev(window[-100:]), 3)

    return {
        "ao5": avg_last(5),
        "ao12": avg_last(12),
        "ao100": avg_last(100),
        "consistency_stdev100": stdev100,
    }


# --------------------------------------------------------------------------- #
# Excel I/O
# --------------------------------------------------------------------------- #

def backup_workbook(excel_path: str, backup_dir: Optional[str], max_backups: int) -> None:
    if not backup_dir:
        backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(excel_path))[0]
    dest = os.path.join(backup_dir, f"{base}_{stamp}.xlsx")
    shutil.copy2(excel_path, dest)
    log.info(f"Backup written to {dest}")

    backups = sorted(glob.glob(os.path.join(backup_dir, f"{base}_*.xlsx")), key=os.path.getmtime)
    for old in backups[:-max_backups]:
        os.remove(old)
        log.info(f"Pruned old backup {old}")


def find_last_filled_row(ws, col: int = 1, start_row: int = 2) -> int:
    row = start_row
    while ws.cell(row=row, column=col).value is not None:
        row += 1
    return row  # first empty row


def append_solves(ws, solves_data: List[Tuple[int, float, str]], start_row: int, new_count: int) -> None:
    """Writes Solve Number / Time / PLL / ao5 / ao12 / ao100 / StdDev(100) for
    the newest `new_count` solves, using the full history for rolling calcs."""
    all_times = [t for _, t, _ in solves_data]
    batch = solves_data[-new_count:]

    row = start_row
    for solve_num, time_sec, pll in batch:
        metrics = rolling_metrics(all_times, solve_num)
        values = [solve_num, time_sec, pll,
                  metrics["ao5"], metrics["ao12"], metrics["ao100"], metrics["consistency_stdev100"]]
        for col_offset, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_offset)
            cell.value = val
            cell.alignment = ALIGNMENT
            if col_offset >= 4 and val is not None:
                cell.number_format = "0.000"
        row += 1
    log.info(f"Appended {new_count} solves (rows {start_row}-{row - 1}), with ao5/ao12/ao100/stdev.")


def change_cell(cell, value, size, bold):
    cell.value = value
    cell.font = Font(name="Aptos Narrow", size=size, bold=bold)
    cell.alignment = ALIGNMENT


def relative_cell(top_left: str, row_offset: int, col_offset: int) -> str:
    row, col = coordinate_to_tuple(top_left)
    return f"{get_column_letter(col + col_offset)}{row + row_offset}"


def add_thick_border(ws, cell_range: str) -> None:
    thick = Side(border_style="thick", color="000000")
    start_cell, end_cell = cell_range.split(":")
    start_row = int("".join(filter(str.isdigit, start_cell)))
    end_row = int("".join(filter(str.isdigit, end_cell)))
    _, start_col = coordinate_to_tuple(start_cell)
    _, end_col = coordinate_to_tuple(end_cell)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                top=thick if row == start_row else cell.border.top,
                bottom=thick if row == end_row else cell.border.bottom,
                left=thick if col == start_col else cell.border.left,
                right=thick if col == end_col else cell.border.right,
            )


def find_col_increment(ws, top_left: str) -> int:
    row, orig_col = coordinate_to_tuple(top_left)
    col = orig_col + 1
    while True:
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and v.lower().endswith("day"):
            return col - orig_col
        col += 1


def find_row_increment(ws, top_left: str) -> int:
    orig_row, col = coordinate_to_tuple(top_left)
    row = orig_row + 1
    while True:
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and v.lower().endswith("day"):
            return row - orig_row
        row += 1


def find_next_summary_block(ws, top_left: str, top_right: str) -> str:
    col_inc = find_col_increment(ws, top_left)
    row_inc = find_row_increment(ws, top_left)
    row, col = coordinate_to_tuple(top_left)

    while ws.cell(row=row, column=col).value is not None:
        row += row_inc
    row -= row_inc
    while ws.cell(row=row, column=col).value is not None:
        col += col_inc

    _, tr_col = coordinate_to_tuple(top_right)
    if tr_col < col:
        _, tl_col = coordinate_to_tuple(top_left)
        col = tl_col
        row += row_inc

    return f"{get_column_letter(col)}{row}"


def write_summary_block(ws, excel_start_row_before_append: int, new_count: int, date_cell: str) -> None:
    """Writes the OVERALL STATS block (bucketed counts, totals, avg/median)
    for the batch of solves just appended."""
    first_data_row = excel_start_row_before_append + 2
    last_data_row = first_data_row + new_count - 1
    value_range = f"B{first_data_row}:B{last_data_row}"

    row, start_col = coordinate_to_tuple(date_cell)
    end_col = start_col + 2
    ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
    ws.merge_cells(f"{get_column_letter(start_col)}{row + 1}:{get_column_letter(end_col)}{row + 1}")

    count_cell = relative_cell(date_cell, 11, 1)
    add_thick_border(ws, f"{date_cell}:{relative_cell(date_cell, 14, 2)}")

    change_cell(ws[date_cell], datetime.today().strftime("%m/%d/%Y/%A"), 12, True)
    change_cell(ws[relative_cell(date_cell, 1, 0)], "OVERALL STATS", 12, True)

    for i in range(6):
        threshold = 15 + i
        change_cell(ws[relative_cell(date_cell, 3 + i, 0)], f"Sub {threshold}", 12, True)
        change_cell(ws[relative_cell(date_cell, 3 + i, 1)],
                    f'=COUNTIF({value_range}, "<={threshold}")', 11, False)

    change_cell(ws[relative_cell(date_cell, 9, 0)], "Above 25", 12, True)
    change_cell(ws[relative_cell(date_cell, 9, 1)], f'=COUNTIF({value_range}, ">=25")', 11, False)

    change_cell(ws[relative_cell(date_cell, 2, 1)], "Solves", 12, True)
    change_cell(ws[relative_cell(date_cell, 2, 2)], "Percent", 12, True)

    for i in range(7):
        pct_cell = relative_cell(date_cell, 3 + i, 2)
        change_cell(ws[pct_cell], f"=({relative_cell(date_cell, 3 + i, 1)}/{count_cell}) * 100", 11, False)
        ws[pct_cell].number_format = "0.00"

    titles = ["Total", "Hours Solving", "Avg Time", "Median Time"]
    formulas = [
        f"=COUNT({value_range})",
        f"=SUM({value_range}) / 60 / 60",
        f"=SUM({value_range})/{count_cell}",
        f"=MEDIAN({value_range})",
    ]
    for i, (title, formula) in enumerate(zip(titles, formulas)):
        change_cell(ws[relative_cell(date_cell, 11 + i, 0)], title, 12, True)
        change_cell(ws[relative_cell(date_cell, 11 + i, 1)], formula, 11, False)

    ws[relative_cell(date_cell, 12, 1)].number_format = "0.000"
    ws[relative_cell(date_cell, 13, 1)].number_format = "0.000"

    log.info(f"Summary block written at {date_cell}.")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(description="Sync cstimer export into Cubing.xlsx with rolling stats.")
    parser.add_argument("--downloads-dir", default=None)
    parser.add_argument("--excel-path", default=None)
    parser.add_argument("--sheet-name", default=None)
    parser.add_argument("--backup-dir", default=None)
    parser.add_argument("--dry-run", action="store_true", help="Parse + compute stats, skip writing/saving.")
    args = parser.parse_args()

    cfg = load_config(args)

    export_path = find_latest_cstimer_export(cfg.downloads_dir)
    solves_data = parse_cstimer_export(export_path)

    if not solves_data:
        log.warning("No solves parsed from export. Nothing to do.")
        return

    if not os.path.exists(cfg.excel_path):
        raise FileNotFoundError(f"Excel workbook not found: {cfg.excel_path}")

    wb = openpyxl.load_workbook(cfg.excel_path)
    ws = wb[cfg.sheet_name]

    next_empty_row = find_last_filled_row(ws)
    latest_excel_solve_num = ws.cell(row=next_empty_row - 1, column=1).value or 0
    latest_parsed_solve_num = solves_data[-1][0]
    new_count = latest_parsed_solve_num - latest_excel_solve_num

    if new_count <= 0:
        log.info("Workbook is already up to date. No new solves to add.")
        return

    log.info(f"{new_count} new solve(s) to add (solve #{latest_excel_solve_num + 1} "
              f"through #{latest_parsed_solve_num}).")

    if args.dry_run:
        preview = rolling_metrics([t for _, t, _ in solves_data], latest_parsed_solve_num)
        log.info(f"[DRY RUN] Would append {new_count} rows. Latest rolling stats: {preview}")
        return

    backup_workbook(cfg.excel_path, cfg.backup_dir, cfg.max_backups)

    append_solves(ws, solves_data, next_empty_row, new_count)
    summary_anchor = find_next_summary_block(ws, "AN8", "BD8")
    write_summary_block(ws, latest_excel_solve_num, new_count, summary_anchor)

    wb.save(cfg.excel_path)
    log.info(f"Saved {cfg.excel_path}.")


if __name__ == "__main__":
    main()